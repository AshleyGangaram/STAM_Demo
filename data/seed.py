"""
STAM Demo Data Seed Script
Run once to:
  1. Create data/demo/projects.xlsx (for the import wizard demo)
  2. Create data/demo/facilities.geojson
  3. Create data/demo/gsdf_zones.geojson
  4. Create data/demo/population.geojson
  5. Seed the SQLite database with all demo data

Usage:
    cd "c:/Data/OneDrive - VITAL TERRA/Applications/STAM/app"
    python data/seed.py
"""

import json
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from services.db import init_db, get_session, Project, Facility, ScoreTemplate, AuditLog

# ── Paths ─────────────────────────────────────────────────────────────────────

DEMO_DIR = os.path.join(os.path.dirname(__file__), "demo")
os.makedirs(DEMO_DIR, exist_ok=True)

# ── Demo projects ─────────────────────────────────────────────────────────────

PROJECTS = [
    {
        "project_id": "P001", "name": "Tembisa Ext 25 Community Clinic",
        "department": "Health", "project_type": "clinic",
        "latitude": -26.002, "longitude": 28.228,
        "budget_rands": 15_000_000, "budget_year": "2026/27",
        "readiness_status": "Ready", "municipality": "Ekurhuleni",
        "ward": "Ward 14", "gsdf_classification": "Priority",
        "total_score": 80, "classification": "Priority Now",
        "score_breakdown": json.dumps({
            "gsdf_overlap": 20, "msdf_priority": 14, "service_gap": 18,
            "transport_access": 8, "brownfield": 8, "readiness": 15, "cost_efficiency": 8,
        }),
    },
    {
        "project_id": "P002", "name": "Alexandra Township Youth Centre",
        "department": "Social Development", "project_type": "community",
        "latitude": -26.102, "longitude": 28.087,
        "budget_rands": 12_000_000, "budget_year": "2026/27",
        "readiness_status": "Ready", "municipality": "City of Johannesburg",
        "ward": "Ward 81", "gsdf_classification": "Priority",
        "total_score": 85, "classification": "Priority Now",
        "score_breakdown": json.dumps({
            "gsdf_overlap": 20, "msdf_priority": 15, "service_gap": 19,
            "transport_access": 9, "brownfield": 9, "readiness": 15, "cost_efficiency": 9,
        }),
    },
    {
        "project_id": "P003", "name": "Soweto Community Clinic — Naledi",
        "department": "Health", "project_type": "clinic",
        "latitude": -26.268, "longitude": 27.869,
        "budget_rands": 18_000_000, "budget_year": "2026/27",
        "readiness_status": "Ready", "municipality": "City of Johannesburg",
        "ward": "Ward 11", "gsdf_classification": "Priority",
        "total_score": 79, "classification": "Priority Next Cycle",
        "score_breakdown": json.dumps({
            "gsdf_overlap": 20, "msdf_priority": 13, "service_gap": 17,
            "transport_access": 8, "brownfield": 8, "readiness": 15, "cost_efficiency": 7,
        }),
    },
    {
        "project_id": "P004", "name": "Diepsloot New Primary School",
        "department": "Education", "project_type": "school",
        "latitude": -26.014, "longitude": 27.958,
        "budget_rands": 22_000_000, "budget_year": "2026/27",
        "readiness_status": "Planning", "municipality": "City of Johannesburg",
        "ward": "Ward 95", "gsdf_classification": "Priority",
        "total_score": 76, "classification": "Priority Next Cycle",
        "score_breakdown": json.dumps({
            "gsdf_overlap": 20, "msdf_priority": 14, "service_gap": 18,
            "transport_access": 7, "brownfield": 7, "readiness": 6, "cost_efficiency": 8,
        }),
    },
    {
        "project_id": "P005", "name": "Katlehong Primary School Extension",
        "department": "Education", "project_type": "school",
        "latitude": -26.349, "longitude": 28.159,
        "budget_rands": 25_000_000, "budget_year": "2026/27",
        "readiness_status": "Planning", "municipality": "Ekurhuleni",
        "ward": "Ward 42", "gsdf_classification": "Priority",
        "total_score": 73, "classification": "Priority Next Cycle",
        "score_breakdown": json.dumps({
            "gsdf_overlap": 20, "msdf_priority": 13, "service_gap": 16,
            "transport_access": 7, "brownfield": 7, "readiness": 6, "cost_efficiency": 7,
        }),
    },
    {
        "project_id": "P006", "name": "Mamelodi West Community Library",
        "department": "Social Development", "project_type": "library",
        "latitude": -25.701, "longitude": 28.378,
        "budget_rands": 8_000_000, "budget_year": "2026/27",
        "readiness_status": "Planning", "municipality": "City of Tshwane",
        "ward": "Ward 95", "gsdf_classification": "Priority",
        "total_score": 71, "classification": "Priority Next Cycle",
        "score_breakdown": json.dumps({
            "gsdf_overlap": 20, "msdf_priority": 12, "service_gap": 16,
            "transport_access": 7, "brownfield": 7, "readiness": 6, "cost_efficiency": 7,
        }),
    },
    {
        "project_id": "P007", "name": "Evaton Community Hall Upgrade",
        "department": "Social Development", "project_type": "community",
        "latitude": -26.529, "longitude": 27.844,
        "budget_rands": 6_000_000, "budget_year": "2026/27",
        "readiness_status": "Planning", "municipality": "Sedibeng",
        "ward": "Ward 7", "gsdf_classification": "Priority",
        "total_score": 70, "classification": "Priority Next Cycle",
        "score_breakdown": json.dumps({
            "gsdf_overlap": 20, "msdf_priority": 12, "service_gap": 15,
            "transport_access": 7, "brownfield": 7, "readiness": 6, "cost_efficiency": 7,
        }),
    },
    {
        "project_id": "P008", "name": "Orange Farm Human Settlements Development",
        "department": "Human Settlements", "project_type": "housing",
        "latitude": -26.477, "longitude": 27.844,
        "budget_rands": 80_000_000, "budget_year": "2026/27",
        "readiness_status": "Concept", "municipality": "City of Johannesburg",
        "ward": "Ward 5", "gsdf_classification": "Priority",
        "total_score": 68, "classification": "Priority Next Cycle",
        "score_breakdown": json.dumps({
            "gsdf_overlap": 20, "msdf_priority": 13, "service_gap": 17,
            "transport_access": 6, "brownfield": 5, "readiness": 2, "cost_efficiency": 8,
        }),
    },
    {
        "project_id": "P009", "name": "Midrand Affordable Housing Phase 2",
        "department": "Human Settlements", "project_type": "housing",
        "latitude": -25.997, "longitude": 28.126,
        "budget_rands": 120_000_000, "budget_year": "2026/27",
        "readiness_status": "Concept", "municipality": "City of Johannesburg",
        "ward": "Ward 103", "gsdf_classification": "Priority",
        "total_score": 64, "classification": "Conditional",
        "score_breakdown": json.dumps({
            "gsdf_overlap": 20, "msdf_priority": 12, "service_gap": 12,
            "transport_access": 7, "brownfield": 7, "readiness": 2, "cost_efficiency": 6,
        }),
    },
    {
        "project_id": "P010", "name": "Atteridgeville Road Upgrade — Phase 1",
        "department": "Roads and Transport", "project_type": "road",
        "latitude": -25.782, "longitude": 27.993,
        "budget_rands": 45_000_000, "budget_year": "2026/27",
        "readiness_status": "Design", "municipality": "City of Tshwane",
        "ward": "Ward 90", "gsdf_classification": "Accommodate",
        "total_score": 55, "classification": "Conditional",
        "score_breakdown": json.dumps({
            "gsdf_overlap": 10, "msdf_priority": 10, "service_gap": 12,
            "transport_access": 8, "brownfield": 7, "readiness": 10, "cost_efficiency": 7,
        }),
    },
    {
        "project_id": "P011", "name": "Centurion Low-Cost Housing Scheme",
        "department": "Human Settlements", "project_type": "housing",
        "latitude": -25.855, "longitude": 28.191,
        "budget_rands": 90_000_000, "budget_year": "2026/27",
        "readiness_status": "Design", "municipality": "City of Tshwane",
        "ward": "Ward 69", "gsdf_classification": "Accommodate",
        "total_score": 52, "classification": "Conditional",
        "score_breakdown": json.dumps({
            "gsdf_overlap": 10, "msdf_priority": 10, "service_gap": 10,
            "transport_access": 7, "brownfield": 5, "readiness": 10, "cost_efficiency": 7,
        }),
    },
    {
        "project_id": "P012", "name": "Germiston Road Upgrade — Industrial Link",
        "department": "Roads and Transport", "project_type": "road",
        "latitude": -26.218, "longitude": 28.165,
        "budget_rands": 60_000_000, "budget_year": "2026/27",
        "readiness_status": "Ready", "municipality": "Ekurhuleni",
        "ward": "Ward 29", "gsdf_classification": "Accommodate",
        "total_score": 48, "classification": "Not Recommended",
        "score_breakdown": json.dumps({
            "gsdf_overlap": 10, "msdf_priority": 8, "service_gap": 8,
            "transport_access": 7, "brownfield": 5, "readiness": 15, "cost_efficiency": 4,
        }),
    },
    {
        "project_id": "P013", "name": "Pretoria CBD Area Clinic",
        "department": "Health", "project_type": "clinic",
        "latitude": -25.745, "longitude": 28.187,
        "budget_rands": 20_000_000, "budget_year": "2026/27",
        "readiness_status": "Concept", "municipality": "City of Tshwane",
        "ward": "Ward 58", "gsdf_classification": "Outside",
        "total_score": 38, "classification": "Not Recommended",
        "score_breakdown": json.dumps({
            "gsdf_overlap": 0, "msdf_priority": 8, "service_gap": 10,
            "transport_access": 8, "brownfield": 6, "readiness": 2, "cost_efficiency": 7,
        }),
    },
    {
        "project_id": "P014", "name": "Sandton Mixed-Use Development",
        "department": "Economic Development", "project_type": "commercial",
        "latitude": -26.107, "longitude": 28.051,
        "budget_rands": 500_000_000, "budget_year": "2026/27",
        "readiness_status": "Concept", "municipality": "City of Johannesburg",
        "ward": "Ward 105", "gsdf_classification": "Outside",
        "total_score": 28, "classification": "Not Recommended",
        "score_breakdown": json.dumps({
            "gsdf_overlap": 0, "msdf_priority": 5, "service_gap": 5,
            "transport_access": 8, "brownfield": 5, "readiness": 2, "cost_efficiency": 3,
        }),
    },
    {
        "project_id": "P015", "name": "Edenvale Industrial Park Expansion",
        "department": "Economic Development", "project_type": "industrial",
        "latitude": -26.121, "longitude": 28.162,
        "budget_rands": 200_000_000, "budget_year": "2026/27",
        "readiness_status": "Concept", "municipality": "Ekurhuleni",
        "ward": "Ward 30", "gsdf_classification": "Discourage",
        "total_score": 22, "classification": "Not Recommended",
        "score_breakdown": json.dumps({
            "gsdf_overlap": 2, "msdf_priority": 3, "service_gap": 5,
            "transport_access": 4, "brownfield": 4, "readiness": 2, "cost_efficiency": 2,
        }),
    },
]

# ── Existing facilities ───────────────────────────────────────────────────────

FACILITIES = [
    # Clinics
    {"name": "Tembisa Main Clinic", "facility_type": "clinic",
     "latitude": -26.018, "longitude": 28.210, "capacity": 400, "current_occupancy": 420,
     "municipality": "Ekurhuleni", "ward": "Ward 12"},
    {"name": "Ivory Park Clinic", "facility_type": "clinic",
     "latitude": -25.985, "longitude": 28.202, "capacity": 300, "current_occupancy": 280,
     "municipality": "Ekurhuleni", "ward": "Ward 9"},
    {"name": "Alexandra Clinic", "facility_type": "clinic",
     "latitude": -26.104, "longitude": 28.095, "capacity": 500, "current_occupancy": 490,
     "municipality": "City of Johannesburg", "ward": "Ward 82"},
    {"name": "Soweto Chiawelo Clinic", "facility_type": "clinic",
     "latitude": -26.285, "longitude": 27.880, "capacity": 350, "current_occupancy": 330,
     "municipality": "City of Johannesburg", "ward": "Ward 10"},
    {"name": "Mamelodi Clinic", "facility_type": "clinic",
     "latitude": -25.710, "longitude": 28.388, "capacity": 450, "current_occupancy": 460,
     "municipality": "City of Tshwane", "ward": "Ward 94"},
    {"name": "Katlehong Clinic", "facility_type": "clinic",
     "latitude": -26.362, "longitude": 28.148, "capacity": 300, "current_occupancy": 320,
     "municipality": "Ekurhuleni", "ward": "Ward 41"},
    {"name": "Diepsloot Clinic", "facility_type": "clinic",
     "latitude": -26.025, "longitude": 27.968, "capacity": 250, "current_occupancy": 310,
     "municipality": "City of Johannesburg", "ward": "Ward 94"},
    {"name": "Orange Farm Clinic", "facility_type": "clinic",
     "latitude": -26.490, "longitude": 27.832, "capacity": 280, "current_occupancy": 350,
     "municipality": "City of Johannesburg", "ward": "Ward 5"},
    # Schools
    {"name": "Tembisa Secondary School", "facility_type": "school",
     "latitude": -26.009, "longitude": 28.220, "capacity": 1200, "current_occupancy": 1350,
     "municipality": "Ekurhuleni", "ward": "Ward 13"},
    {"name": "Diepsloot Secondary", "facility_type": "school",
     "latitude": -26.008, "longitude": 27.965, "capacity": 800, "current_occupancy": 1050,
     "municipality": "City of Johannesburg", "ward": "Ward 95"},
    {"name": "Katlehong High School", "facility_type": "school",
     "latitude": -26.344, "longitude": 28.165, "capacity": 1100, "current_occupancy": 1380,
     "municipality": "Ekurhuleni", "ward": "Ward 43"},
    {"name": "Mamelodi East Secondary", "facility_type": "school",
     "latitude": -25.695, "longitude": 28.392, "capacity": 900, "current_occupancy": 870,
     "municipality": "City of Tshwane", "ward": "Ward 96"},
    {"name": "Soweto High School Naledi", "facility_type": "school",
     "latitude": -26.279, "longitude": 27.862, "capacity": 1000, "current_occupancy": 980,
     "municipality": "City of Johannesburg", "ward": "Ward 12"},
    {"name": "Alexandra Secondary School", "facility_type": "school",
     "latitude": -26.109, "longitude": 28.080, "capacity": 1100, "current_occupancy": 1250,
     "municipality": "City of Johannesburg", "ward": "Ward 80"},
    # Libraries
    {"name": "Tembisa Public Library", "facility_type": "library",
     "latitude": -26.012, "longitude": 28.213, "capacity": 200, "current_occupancy": 120,
     "municipality": "Ekurhuleni", "ward": "Ward 13"},
    {"name": "Alexandra Public Library", "facility_type": "library",
     "latitude": -26.100, "longitude": 28.089, "capacity": 150, "current_occupancy": 90,
     "municipality": "City of Johannesburg", "ward": "Ward 81"},
    {"name": "Soweto Meadowlands Library", "facility_type": "library",
     "latitude": -26.255, "longitude": 27.895, "capacity": 180, "current_occupancy": 110,
     "municipality": "City of Johannesburg", "ward": "Ward 16"},
    # Community Halls
    {"name": "Diepsloot Community Hall", "facility_type": "community_hall",
     "latitude": -26.021, "longitude": 27.961, "capacity": 500, "current_occupancy": 250,
     "municipality": "City of Johannesburg", "ward": "Ward 94"},
    {"name": "Orange Farm Community Hall", "facility_type": "community_hall",
     "latitude": -26.482, "longitude": 27.838, "capacity": 400, "current_occupancy": 320,
     "municipality": "City of Johannesburg", "ward": "Ward 4"},
    {"name": "Evaton Community Hall", "facility_type": "community_hall",
     "latitude": -26.525, "longitude": 27.850, "capacity": 350, "current_occupancy": 200,
     "municipality": "Sedibeng", "ward": "Ward 8"},
]


def _make_point(lat, lon, props):
    return {
        "type": "Feature",
        "geometry": {"type": "Point", "coordinates": [lon, lat]},
        "properties": props,
    }


def _make_polygon(coords, props):
    return {
        "type": "Feature",
        "geometry": {"type": "Polygon", "coordinates": [coords]},
        "properties": props,
    }


def write_facilities_geojson():
    features = []
    for f in FACILITIES:
        features.append(_make_point(f["latitude"], f["longitude"], {
            "name": f["name"],
            "facility_type": f["facility_type"],
            "capacity": f["capacity"],
            "current_occupancy": f["current_occupancy"],
            "municipality": f["municipality"],
            "ward": f["ward"],
            "occupancy_pct": round(f["current_occupancy"] / f["capacity"] * 100, 1),
        }))
    geojson = {"type": "FeatureCollection", "features": features}
    path = os.path.join(DEMO_DIR, "facilities.geojson")
    with open(path, "w") as fp:
        json.dump(geojson, fp, indent=2)
    print(f"  wrote {path} ({len(features)} features)")


def write_gsdf_zones_geojson():
    """Simplified GSDF 2030 zones for Gauteng demo."""
    features = [
        # Priority zones — dense urban corridors and townships
        _make_polygon([
            [27.90, -26.30], [28.00, -26.30], [28.00, -26.15], [27.90, -26.15], [27.90, -26.30]
        ], {"zone_name": "Soweto Spatial Priority Corridor", "classification": "Priority",
            "municipality": "City of Johannesburg"}),
        _make_polygon([
            [28.05, -26.15], [28.20, -26.15], [28.20, -25.98], [28.05, -25.98], [28.05, -26.15]
        ], {"zone_name": "Alexandra-Tembisa Priority Zone", "classification": "Priority",
            "municipality": "City of Johannesburg / Ekurhuleni"}),
        _make_polygon([
            [27.82, -26.55], [27.95, -26.55], [27.95, -26.42], [27.82, -26.42], [27.82, -26.55]
        ], {"zone_name": "Evaton-Orange Farm Priority Zone", "classification": "Priority",
            "municipality": "Sedibeng / City of Johannesburg"}),
        _make_polygon([
            [28.12, -26.40], [28.22, -26.40], [28.22, -26.28], [28.12, -26.28], [28.12, -26.40]
        ], {"zone_name": "Katlehong Priority Zone", "classification": "Priority",
            "municipality": "Ekurhuleni"}),
        _make_polygon([
            [28.30, -25.75], [28.45, -25.75], [28.45, -25.64], [28.30, -25.64], [28.30, -25.75]
        ], {"zone_name": "Mamelodi Priority Zone", "classification": "Priority",
            "municipality": "City of Tshwane"}),
        _make_polygon([
            [27.90, -26.10], [28.05, -26.10], [28.05, -25.95], [27.90, -25.95], [27.90, -26.10]
        ], {"zone_name": "Diepsloot Priority Zone", "classification": "Priority",
            "municipality": "City of Johannesburg"}),
        # Accommodate zones
        _make_polygon([
            [27.92, -25.90], [28.22, -25.90], [28.22, -25.78], [27.92, -25.78], [27.92, -25.90]
        ], {"zone_name": "Northern Tshwane Accommodate Zone", "classification": "Accommodate",
            "municipality": "City of Tshwane"}),
        _make_polygon([
            [28.10, -25.95], [28.28, -25.95], [28.28, -25.82], [28.10, -25.82], [28.10, -25.95]
        ], {"zone_name": "Midrand-Centurion Accommodate Zone", "classification": "Accommodate",
            "municipality": "City of Johannesburg / City of Tshwane"}),
        # Discourage zone — peripheral low-density / industrial
        _make_polygon([
            [28.10, -26.18], [28.25, -26.18], [28.25, -26.08], [28.10, -26.08], [28.10, -26.18]
        ], {"zone_name": "Edenvale-Bedfordview Industrial Discourage Zone",
            "classification": "Discourage", "municipality": "Ekurhuleni"}),
    ]
    geojson = {"type": "FeatureCollection", "features": features}
    path = os.path.join(DEMO_DIR, "gsdf_zones.geojson")
    with open(path, "w") as fp:
        json.dump(geojson, fp, indent=2)
    print(f"  wrote {path} ({len(features)} features)")


def write_population_geojson():
    """Simplified population density polygons (Gauteng ward approximations)."""
    features = [
        _make_polygon([
            [28.05, -26.05], [28.25, -26.05], [28.25, -25.95], [28.05, -25.95], [28.05, -26.05]
        ], {"ward_name": "Tembisa Wards 12-15", "municipality": "Ekurhuleni",
            "population": 185000, "area_km2": 22, "density_per_km2": 8409}),
        _make_polygon([
            [28.06, -26.15], [28.12, -26.15], [28.12, -26.08], [28.06, -26.08], [28.06, -26.15]
        ], {"ward_name": "Alexandra Wards 79-82", "municipality": "City of Johannesburg",
            "population": 75000, "area_km2": 5, "density_per_km2": 15000}),
        _make_polygon([
            [27.83, -26.35], [27.98, -26.35], [27.98, -26.22], [27.83, -26.22], [27.83, -26.35]
        ], {"ward_name": "Soweto North Wards 10-16", "municipality": "City of Johannesburg",
            "population": 210000, "area_km2": 42, "density_per_km2": 5000}),
        _make_polygon([
            [27.90, -26.05], [28.02, -26.05], [28.02, -25.97], [27.90, -25.97], [27.90, -26.05]
        ], {"ward_name": "Diepsloot Wards 94-95", "municipality": "City of Johannesburg",
            "population": 95000, "area_km2": 10, "density_per_km2": 9500}),
        _make_polygon([
            [28.12, -26.40], [28.22, -26.40], [28.22, -26.28], [28.12, -26.28], [28.12, -26.40]
        ], {"ward_name": "Katlehong Wards 40-44", "municipality": "Ekurhuleni",
            "population": 140000, "area_km2": 17, "density_per_km2": 8235}),
        _make_polygon([
            [28.32, -25.75], [28.43, -25.75], [28.43, -25.65], [28.32, -25.65], [28.32, -25.75]
        ], {"ward_name": "Mamelodi Wards 93-97", "municipality": "City of Tshwane",
            "population": 120000, "area_km2": 25, "density_per_km2": 4800}),
        _make_polygon([
            [27.82, -26.55], [27.92, -26.55], [27.92, -26.44], [27.82, -26.44], [27.82, -26.55]
        ], {"ward_name": "Evaton-Orange Farm Wards 3-8", "municipality": "Sedibeng / CoJ",
            "population": 160000, "area_km2": 30, "density_per_km2": 5333}),
        _make_polygon([
            [28.10, -25.98], [28.18, -25.98], [28.18, -25.91], [28.10, -25.91], [28.10, -25.98]
        ], {"ward_name": "Midrand Wards 100-104", "municipality": "City of Johannesburg",
            "population": 85000, "area_km2": 32, "density_per_km2": 2656}),
        _make_polygon([
            [27.95, -25.82], [28.02, -25.82], [28.02, -25.76], [27.95, -25.76], [27.95, -25.82]
        ], {"ward_name": "Atteridgeville Wards 88-92", "municipality": "City of Tshwane",
            "population": 65000, "area_km2": 18, "density_per_km2": 3611}),
    ]
    geojson = {"type": "FeatureCollection", "features": features}
    path = os.path.join(DEMO_DIR, "population.geojson")
    with open(path, "w") as fp:
        json.dump(geojson, fp, indent=2)
    print(f"  wrote {path} ({len(features)} features)")


def write_projects_xlsx():
    """Write projects.xlsx with intentional validation errors for the import demo."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Capital Projects"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    error_fill = PatternFill("solid", fgColor="FFE0E0")

    headers = [
        "project_id", "project_name", "department", "project_type",
        "latitude", "longitude", "budget_rands", "budget_year",
        "readiness_status", "municipality", "ward",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 13 clean rows (P001–P013 minus intentional errors)
    for p in PROJECTS[:-2]:
        ws.append([
            p["project_id"], p["name"], p["department"], p["project_type"],
            p["latitude"], p["longitude"], p["budget_rands"], p["budget_year"],
            p["readiness_status"], p["municipality"], p["ward"],
        ])

    # Row with invalid budget (demo validation error — P014)
    p = PROJECTS[13]
    ws.append([
        p["project_id"], p["name"], p["department"], p["project_type"],
        p["latitude"], p["longitude"], "FIVE_HUNDRED_MILLION",   # <-- invalid budget
        p["budget_year"], p["readiness_status"], p["municipality"], p["ward"],
    ])
    # Highlight the error row
    err_row = ws.max_row
    for cell in ws[err_row]:
        cell.fill = error_fill

    # Row with missing coordinates (P015)
    p = PROJECTS[14]
    ws.append([
        p["project_id"], p["name"], p["department"], p["project_type"],
        "", "",   # <-- missing coordinates
        p["budget_rands"], p["budget_year"],
        p["readiness_status"], p["municipality"], p["ward"],
    ])
    err_row = ws.max_row
    for cell in ws[err_row]:
        cell.fill = error_fill

    # Column widths
    for i, col in enumerate(["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"], 1):
        widths = [12, 40, 22, 14, 10, 10, 20, 12, 16, 25, 12]
        ws.column_dimensions[col].width = widths[i - 1]

    path = os.path.join(DEMO_DIR, "projects.xlsx")
    wb.save(path)
    print(f"  wrote {path} ({len(PROJECTS)} rows including 2 intentional errors)")


def seed_database():
    engine = init_db()
    session = get_session(engine)

    # Clear existing demo data
    session.query(AuditLog).delete()
    session.query(Project).delete()
    session.query(Facility).delete()
    session.query(ScoreTemplate).delete()
    session.commit()

    # Seed projects
    for p in PROJECTS:
        proj = Project(**{k: v for k, v in p.items()})
        session.add(proj)

    # Seed facilities
    for f in FACILITIES:
        fac = Facility(**f)
        session.add(fac)

    # Seed default score template
    import json as _json
    default_weights = {
        "gsdf_overlap": 20, "msdf_priority": 15, "service_gap": 20,
        "transport_access": 10, "brownfield": 10, "readiness": 15, "cost_efficiency": 10,
    }
    template = ScoreTemplate(
        template_name="Default GSDF 2030 Weights",
        weights=_json.dumps(default_weights),
        active=1,
    )
    session.add(template)

    # Seed initial audit entry
    entry = AuditLog(
        user_role="System",
        action="DATABASE_SEEDED",
        entity_type="system",
        detail=_json.dumps({"projects": len(PROJECTS), "facilities": len(FACILITIES)}),
    )
    session.add(entry)

    session.commit()
    session.close()
    print(f"  seeded DB: {len(PROJECTS)} projects, {len(FACILITIES)} facilities, 1 score template")


if __name__ == "__main__":
    print("STAM Demo Data Seed")
    print("=" * 40)
    print("Writing GeoJSON files...")
    write_facilities_geojson()
    write_gsdf_zones_geojson()
    write_population_geojson()
    print("Writing projects.xlsx...")
    write_projects_xlsx()
    print("Seeding database...")
    seed_database()
    print()
    print("Done. Run: streamlit run app.py")
